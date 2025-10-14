import pandas as pd
import os
import re
from pathlib import Path
from collections import defaultdict

def parse_classes(classes_str):
    """
    Parse the Classes column to extract activity types and their time ranges.
    Example: "Sitting[0 to 1]; Sleeping[2 to 8]" -> {'Sitting': '0:1', 'Sleeping': '2:8'}
    """
    if pd.isna(classes_str) or not classes_str.strip():
        return {}
    
    activities = {}
    # Split by semicolon to get individual activities
    parts = classes_str.split(';')
    
    for part in parts:
        part = part.strip()
        # Match pattern: ActivityName[time range]
        match = re.match(r'([A-Za-z\s]+)\[(.*?)\]', part)
        if match:
            activity_name = match.group(1).strip()
            time_range = match.group(2).strip()
            # Convert "0 to 1" to "0:1"
            time_range = time_range.replace(' to ', ':')
            activities[activity_name] = time_range
    
    return activities

def collect_all_data(base_dir):
    """
    Collect all data from ADL.csv and Fall.csv files across all subjects.
    Returns a structured dictionary.
    """
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    # data structure: data[subject][category][activity] = list of time_ranges
    
    base_path = Path(base_dir)
    
    # Find all subject directories
    subject_dirs = sorted([d for d in base_path.iterdir() if d.is_dir() and d.name.startswith('Subject')])
    
    for subject_dir in subject_dirs:
        subject_name = subject_dir.name
        
        # Process ADL.csv
        adl_csv = subject_dir / 'ADL.csv'
        if adl_csv.exists():
            try:
                df = pd.read_csv(adl_csv)
                for _, row in df.iterrows():
                    classes = row.get('Classes', row.get(' Classes', ''))
                    
                    activities = parse_classes(str(classes))
                    for activity, time_range in activities.items():
                        data[subject_name]['ADL'][activity].append(time_range)
            except Exception as e:
                print(f"Error reading {adl_csv}: {e}")
        
        # Process Fall.csv
        fall_csv = subject_dir / 'Fall.csv'
        if fall_csv.exists():
            try:
                df = pd.read_csv(fall_csv)
                for _, row in df.iterrows():
                    classes = row.get('Classes', row.get(' Classes', ''))
                    
                    activities = parse_classes(str(classes))
                    for activity, time_range in activities.items():
                        data[subject_name]['Fall'][activity].append(time_range)
            except Exception as e:
                print(f"Error reading {fall_csv}: {e}")
    
    return data

def create_wide_format_table(data):
    """
    Create a table with each trial in its own sub-column.
    """
    subjects = sorted(data.keys())
    
    # Collect all unique activities per category
    all_activities = set()
    for subject in subjects:
        for category in ['ADL', 'Fall']:
            all_activities.update(data[subject][category].keys())
    
    all_activities = sorted(all_activities)
    
    # Build column structure
    columns = ['Category', 'Activity']
    column_mapping = []  # Track which subject and trial each column belongs to
    
    for subject in subjects:
        # Find max trials needed for this subject across all activities
        max_subject_trials = 0
        for category in ['ADL', 'Fall']:
            for activity in all_activities:
                num_trials = len(data[subject][category].get(activity, []))
                max_subject_trials = max(max_subject_trials, num_trials)
        
        # Create sub-columns for each trial
        for i in range(1, max_subject_trials + 1):
            columns.append(f"{subject}_Trial{i}")
            column_mapping.append((subject, i, max_subject_trials))
    
    rows = []
    
    # ADL section
    rows.append(['ADL', 'ACTIVITIES OF DAILY LIVING'] + [''] * (len(columns) - 2))
    
    adl_activities = sorted([act for act in all_activities 
                             if any(act in data[s]['ADL'] for s in subjects)])
    
    for activity in adl_activities:
        row = ['ADL', f"  {activity}"]
        for subject in subjects:
            time_ranges = data[subject]['ADL'].get(activity, [])
            # Find how many trial columns this subject has
            max_subject_trials = 0
            for cat in ['ADL', 'Fall']:
                for act in all_activities:
                    num_trials = len(data[subject][cat].get(act, []))
                    max_subject_trials = max(max_subject_trials, num_trials)
            
            # Add each trial to its own column
            for i in range(max_subject_trials):
                if i < len(time_ranges):
                    row.append(time_ranges[i])
                else:
                    row.append('')
        rows.append(row)
    
    # Fall section
    rows.append(['Fall', 'FALL EVENTS'] + [''] * (len(columns) - 2))
    
    fall_activities = sorted([act for act in all_activities 
                              if any(act in data[s]['Fall'] for s in subjects)])
    
    for activity in fall_activities:
        row = ['Fall', f"  {activity}"]
        for subject in subjects:
            time_ranges = data[subject]['Fall'].get(activity, [])
            max_subject_trials = 0
            for cat in ['ADL', 'Fall']:
                for act in all_activities:
                    num_trials = len(data[subject][cat].get(act, []))
                    max_subject_trials = max(max_subject_trials, num_trials)
            
            for i in range(max_subject_trials):
                if i < len(time_ranges):
                    row.append(time_ranges[i])
                else:
                    row.append('')
        rows.append(row)
    
    df = pd.DataFrame(rows, columns=columns)
    return df, column_mapping

def format_excel_with_headers(writer, sheet_name, df, subjects, max_trials_per_subject):
    """
    Add formatted headers with merged cells for subjects.
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    worksheet = writer.sheets[sheet_name]
    
    # Insert a new row at the top for subject headers
    worksheet.insert_rows(1)
    
    # Format the category and activity headers
    worksheet.cell(1, 1, "Category")
    worksheet.cell(1, 1).font = Font(bold=True)
    worksheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet.cell(1, 2, "Activity")
    worksheet.cell(1, 2).font = Font(bold=True)
    worksheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    
    # Merge Category and Activity cells vertically
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    
    # Add subject headers
    col_idx = 3  # Start after Category and Activity
    for subject in subjects:
        num_trials = max_trials_per_subject[subject]
        
        # Merge cells for subject name
        if num_trials > 1:
            worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + num_trials - 1)
        
        # Set subject header
        cell = worksheet.cell(1, col_idx, subject)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Set trial headers
        for i in range(num_trials):
            trial_cell = worksheet.cell(2, col_idx + i, f"Trial {i+1}")
            trial_cell.font = Font(bold=True, size=10)
            trial_cell.alignment = Alignment(horizontal='center', vertical='center')
            trial_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        col_idx += num_trials
    
    # Auto-adjust column widths - iterate by column index instead of cells
    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row_idx, col_idx)
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width
        adjusted_width = min(max(max_length + 2, 10), 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze top two rows
    worksheet.freeze_panes = 'A3'

def main():
    # Set your dataset directory path
    base_dir = "./GMDCSA24-A-Dataset-for-Human-Fall-Detection-in-Videos"
    
    # Or use current directory if script is run from dataset folder
    if not os.path.exists(base_dir):
        base_dir = "."
    
    print("Collecting data from all CSV files...")
    data = collect_all_data(base_dir)
    
    subjects = sorted(data.keys())
    
    # Find max trials per subject for formatting
    all_activities = set()
    for subject in subjects:
        for category in ['ADL', 'Fall']:
            all_activities.update(data[subject][category].keys())
    
    max_trials_per_subject = {}
    for subject in subjects:
        max_trials = 0
        for category in ['ADL', 'Fall']:
            for activity in all_activities:
                num_trials = len(data[subject][category].get(activity, []))
                max_trials = max(max_trials, num_trials)
        max_trials_per_subject[subject] = max_trials
    
    print("Creating wide format table (each trial in separate column)...")
    wide_df, column_mapping = create_wide_format_table(data)
    
    # Save to files
    print("\nSaving tables...")
    
    # Save as Excel with custom formatting
    with pd.ExcelWriter('video_annotations_trials_separate.xlsx', engine='openpyxl') as writer:
        # Write the dataframe
        wide_df.to_excel(writer, sheet_name='Activity Trials', index=False)
        
        # Format with merged headers
        format_excel_with_headers(writer, 'Activity Trials', wide_df, subjects, max_trials_per_subject)
    
    # Save CSV file
    wide_df.to_csv('video_trials_all_data.csv', index=False)
    
    # Save as formatted text
    with open('video_trials_formatted.txt', 'w') as f:
        f.write("="*150 + "\n")
        f.write("VIDEO ACTIVITY TRIALS: Time indices for each activity\n")
        f.write("Format: start:end (in seconds)\n")
        f.write("="*150 + "\n\n")
        f.write(wide_df.to_string(index=False, max_colwidth=20))
    
    print("\n✓ Files created:")
    print("  - video_annotations_trials_separate.xlsx (Excel with formatted headers)")
    print("  - video_trials_all_data.csv (CSV format)")
    print("  - video_trials_formatted.txt (formatted text)")
    
    print("\n" + "="*100)
    print("PREVIEW (first 15 rows):")
    print("="*100)
    print(wide_df.head(15).to_string(index=False, max_colwidth=12))
    
    print("\n\nColumn structure:")
    print(f"Total columns: {len(wide_df.columns)}")
    for subject, count in max_trials_per_subject.items():
        print(f"  {subject}: {count} trial columns")
    
    print("\n✓ Excel file has:")
    print("  - Merged headers showing subject names")
    print("  - Sub-columns for each trial")
    print("  - Frozen header rows for easy scrolling")
    print("  - Auto-sized columns")

if __name__ == "__main__":
    main()

